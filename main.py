import json
import base64
import io
import os
from flask import Flask, request, Response

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

app = Flask(__name__)

@app.route("/", methods=["GET"])
def health():
    return {"status": "ok", "service": "generate-xls"}

@app.route("/generate-xls", methods=["POST", "OPTIONS"])
def generate_xls():
    if request.method == "OPTIONS":
        return _cors(Response("", 200))

    body    = request.get_json(force=True) or {}
    header  = body.get("header", {})
    entries = body.get("entries", [])
    tmpl_b64 = body.get("template_b64", "")

    if not tmpl_b64:
        return _cors(Response(
            json.dumps({"error": "template_b64 is required"}),
            500, content_type="application/json"
        ))

    try:
        wb = load_workbook(io.BytesIO(base64.b64decode(tmpl_b64)))
        ws = wb.active
        ws.title = (header.get("judulDokumen") or "Sheet1").strip()[:31]

        ws["D4"] = header.get("keterangan", "")
        ws["D5"] = header.get("aplikasi", "")
        ws["D6"] = header.get("modul", "")
        ws["D7"] = header.get("createdBy", "")
        ws["D8"] = header.get("testedBy", "")
        ws["D9"] = header.get("targetFinish", "")
        ws["B12"] = header.get("judulDokumen", "")
        ws["B12"].font = Font(bold=True, size=13, name="Calibri")
        ws["B12"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        ROW_H_PT = 14.4
        COL_D_PX = 788
        ROW_H_PX = ROW_H_PT * 96 / 72
        thin = Side(style="thin", color="000000")

        def lr(c): c.border = Border(left=thin, right=thin)
        def lr_b(c): c.border = Border(left=thin, right=thin, bottom=thin)
        def f11(c, bold=False, align="general", color="000000"):
            c.font = Font(bold=bold, size=11, name="Calibri", color=color)
            c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        def sfont(s):
            if s == "NOT OK": return Font(bold=True, size=11, name="Calibri", color="FF0000")
            return Font(bold=(s == "OK"), size=11, name="Calibri", color="000000")
        def emb(url, row):
            if not url or "base64," not in url: return row
            try:
                b64 = url.split("base64,")[1]
                pil = PILImage.open(io.BytesIO(base64.b64decode(b64)))
                ow, oh = pil.size
                dh = int(oh * COL_D_PX / ow)
                nr = max(1, int(dh / ROW_H_PX) + 1)
                for r in range(row, row + nr):
                    ws.row_dimensions[r].height = ROW_H_PT
                buf = io.BytesIO()
                pil.save(buf, format="PNG")
                buf.seek(0)
                xl = XLImage(buf)
                xl.width = COL_D_PX
                xl.height = dh
                ws.add_image(xl, f"D{row}")
                return row + nr
            except:
                return row + 1

        cur = 26
        for ent in entries:
            st = ent.get("status", "")
            tgl = ent.get("tanggalTest", "")
            ws.row_dimensions[cur].height = ROW_H_PT
            ws[f"B{cur}"] = ent.get("no", "")
            ws[f"C{cur}"] = ent.get("object", "")
            ws[f"D{cur}"] = ent.get("keterangan", "")
            ws[f"E{cur}"] = tgl
            ws[f"F{cur}"] = st
            f11(ws[f"B{cur}"], align="center")
            f11(ws[f"C{cur}"], bold=True)
            ws[f"D{cur}"].font = Font(size=11, name="Calibri")
            ws[f"D{cur}"].alignment = Alignment(wrap_text=True, vertical="center")
            f11(ws[f"E{cur}"], align="center")
            ws[f"F{cur}"].font = sfont(st)
            ws[f"F{cur}"].alignment = Alignment(horizontal="center", vertical="center")
            for col in ["B","C","D","E","F"]: lr(ws[f"{col}{cur}"])
            cur += 1

            for img in ent.get("images", []):
                cur = emb(img.get("dataUrl", ""), cur)

            for ss in ent.get("subSections", []):
                if ss.get("deskripsi"):
                    ws.row_dimensions[cur].height = ROW_H_PT
                    ws[f"D{cur}"] = ss["deskripsi"]
                    ws[f"D{cur}"].font = Font(size=11, name="Calibri")
                    ws[f"D{cur}"].alignment = Alignment(wrap_text=True, vertical="center")
                    for col in ["B","C","D","E","F"]: lr(ws[f"{col}{cur}"])
                    cur += 1
                for img in ss.get("images", []):
                    cur = emb(img.get("dataUrl", ""), cur)

            if ent.get("subKeterangan"):
                ws.row_dimensions[cur].height = ROW_H_PT
                ws[f"D{cur}"] = ent["subKeterangan"]
                ws[f"E{cur}"] = tgl
                ws[f"D{cur}"].font = Font(size=11, name="Calibri")
                ws[f"D{cur}"].alignment = Alignment(vertical="center")
                ws[f"E{cur}"].font = Font(size=11, name="Calibri")
                ws[f"E{cur}"].alignment = Alignment(horizontal="center", vertical="center")
                for col in ["B","C","D","E","F"]: lr(ws[f"{col}{cur}"])
                cur += 1

            for col in ["B","C","D","E","F"]: lr_b(ws[f"{col}{cur - 1}"])

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        xlsx = out.read()

        fname = (header.get("judulDokumen") or "scenario-test").strip()
        for ch in r'\/:*?"<>|': fname = fname.replace(ch, "_")

        resp = Response(
            xlsx,
            200,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        resp.headers["Content-Disposition"] = f'attachment; filename="{fname}.xlsx"'
        return _cors(resp)

    except Exception as e:
        return _cors(Response(
            json.dumps({"error": str(e)}),
            500, content_type="application/json"
        ))


def _cors(resp: Response) -> Response:
    resp.headers["Access-Control-Allow-Origin"]  = "*"
    resp.headers["Access-Control-Allow-Headers"] = "Content-Type"
    resp.headers["Access-Control-Allow-Methods"] = "POST, OPTIONS"
    return resp


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8000))
    app.run(host="0.0.0.0", port=port)
